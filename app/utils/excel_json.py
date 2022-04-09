import openpyxl
import json
from openpyxl import Workbook
from openpyxl import load_workbook


def excel_To_Json(path):
    wb = load_workbook(path)

    ws = wb.active

    num_row = ws.max_row
    num_col = ws.max_column
    # print(num_col)
    # print(num_row)
    keys = []
    for i in range(1, num_col + 1):
        keys.append(ws.cell(1, i).value)
    # print(keys)

    change_json = []
    values = []
    for i in range(2, num_row + 1):
        for j in range(1, num_col + 1):
            # print(index[j-1])
            # print(ws.cell(i, j).value)
            # jj = "'%s':%d" %(index[j-1], ws.cell(i, j).value)
            # jj = dict(jj_index=ws.cell(i, j).value)
            values.append(ws.cell(i, j).value)
        change_json_rows = dict(zip(keys, values))
        if change_json_rows:
            change_json.append(change_json_rows)
        change_json_rows = []
        values = []

    # for i in change_json:
    #     data1 = json.dumps(i)  # 列表
    #     print(data1, type(data1))

    json_result = json.dumps(change_json, ensure_ascii=False)  # 列表
    # print(data1, type(data1))

    return change_json

